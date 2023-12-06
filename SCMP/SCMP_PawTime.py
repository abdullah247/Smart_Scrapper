import math
import openpyxl
import time




classParColumns={
    "sh1000M":3,
    "sh1200M":11,
    "sh1400M":19,
    "sh1600M":27,
    "sh1800M":35,
    "sh2000M": 42,
    "sh2200M": 49,
    "sh2400M": 55,
    "hv1000M":66,
    "hv1200M":73,
    "hv1650M":81,
    "hv1800M":87,
    "hv2200M": 94,
    "hv2400M": 100,
    "awt1200M": 110,
    "awt1650M": 116,
    "awt1800M": 122



}

pawratings={}

def class_Name_Switcher(argument):
    switcher = {
        "Group/Listed": 0,
        "1": 1,
        "2": 2,
        "3": 3,
        "4": 4,
        "5": 5,
        "Griffin": 6
    }

    # get() method of dictionary data type returns
    # value of passed argument if it is present
    # in dictionary otherwise second argument will
    # be assigned as default value of passed argument
    return switcher.get(argument, -1)

def row_Switcher(argument):
    switcher = {
        "A": 6,
        "A+3": 8,
        "B": 10,
        "B+2": 12,
        "C": 14,
        "C+3": 16,
        "": 3
    }

    # get() method of dictionary data type returns
    # value of passed argument if it is present
    # in dictionary otherwise second argument will
    # be assigned as default value of passed argument
    return switcher.get(argument, 4)

parheader = {"sv": 3, "hv": 65}
pawtime = {

}
partimefull = {

}
pawCounter = {

}


def getTurnValue(turn, param, val):
    flag = 2
    try:
        myval = float("".join(param.rsplit(".", 1)))

        for i in range(9):
            # print(turn.cell(1,flag).value ,f"{val}m",turn.cell(1,flag).value == f"{val}m")
            if turn.cell(1, flag).value == f"{val}m":
                break
            flag = flag + 3

        current = turn.cell(2, flag).value

        for i in range(2, turn.max_row):
            val = turn.cell(i, flag - 1).value

            comp = float("".join(val.rsplit(".", 1)))

            # print(myval ,comp,myval ==comp,myval>comp)
            if myval == comp or comp > myval:
                current = turn.cell(i, flag).value
                return current
    except:
        return ""

    pass


def saveme(wk, name):
    flag = True
    counter = 0
    print("Updating Latest File Donot Close Or interfere")
    while flag:
        try:
            if counter == 0:
                counter += 1
                wk.save(f"{name}.xlsx")
                flag = False
            else:
                counter += 1
                wk.save(f"{name}{counter}.xlsx")
                flag = False

        except:
            print("Cannot Save File if already opened saving it by adding number at the end")
    print("Updated Latest File")

    pass




# saveme(wk, "BackupPawTime")


def addpawtime(param, param1):
    if "." in param1:
        s2 = param1.split(".")
        mili = 0
        if len(s2) == 3:
            mili = int(s2[2]) + int(s2[1]) * 100 + int(s2[0]) * 60 * 100
            return param + mili





        else:
            return param
    else:

        return param
    pass

def getSCMPPAWTIME(Address):

    wk = openpyxl.load_workbook(Address)
    sp = wk["class par"]
    sh = wk["Database"]
    st = wk["st"]
    hv = wk["hv"]
    stAwt = wk["st awt"]
    for i in range(2, sh.max_row + 1):
        sh.cell(i, 12).value=str(sh.cell(i, 12).value).replace(":",".")
        key = str(sh.cell(i, 15).value) + "," + str(sh.cell(i, 16).value) + "," + str(
            sh.cell(i, 17).value) + "," + str(sh.cell(i, 18).value) + "," + str(sh.cell(i, 20).value).replace("G1","Group/Listed").replace("G2","Group/Listed").replace("G3","Group/Listed").replace("Gri","Griffin")
        key2=str(sh.cell(i, 15).value) + "," + str(sh.cell(i, 16).value) + "," + str(sh.cell(i, 18).value) + "," + str(sh.cell(i, 20).value).replace("G1","Group/Listed").replace("G2","Group/Listed").replace("G3","Group/Listed").replace("Gri","Griffin")
        if not sh.cell(i, 21).value is None and len(str(sh.cell(i, 21).value).split("-")) > 2 and str(sh.cell(i,1).value).strip() =="1":
            try:
                dat = int(str(sh.cell(i, 21).value).split("-")[0].replace("20", "", 1))

                currentYear = int(time.strftime("%y", time.localtime()))
                if dat + 3 >= currentYear:
                    print("G",str(sh.cell(i, 12).value))
                    if key in pawtime:
                        pawtime[key] = addpawtime(pawtime[key], str(sh.cell(i, 12).value))
                        pawCounter[key] += 1
                    else:
                        if "." in str(sh.cell(i, 12).value):

                            stval = str(sh.cell(i, 12).value).split(".")
                            pawCounter[key] = 1
                            if len(stval) == 3:
                                pawtime[key] = int(stval[2]) + int(stval[1]) * 100 + int(stval[0]) * 60 * 100
                            else:
                                pawtime[key] = 0
                                pawCounter[key] = 0
                        else:
                            pawtime[key] = 0
                            pawCounter[key] = 0
                    if key2 in pawtime:
                        pawtime[key2] = addpawtime(pawtime[key2], str(sh.cell(i, 12).value))
                        pawCounter[key2] += 1
                    else:
                        if "." in str(sh.cell(i, 12).value):

                            stval = str(sh.cell(i, 12).value).split(".")
                            pawCounter[key2] = 1
                            if len(stval) == 3:
                                pawtime[key2] = int(stval[2]) + int(stval[1]) * 100 + int(stval[0]) * 60 * 100
                            else:
                                pawtime[key2] = 0
                                pawCounter[key2] =0




            except Exception as e:
                print(e, i)

    for i in range(2, sh.max_row + 1):
        if i % 10000 == 0:
            print(f"Adding Paw Values {i}")
        if not sh.cell(i, 21).value is None and len(str(sh.cell(i, 21).value).split("-")) > 2:

            key = str(sh.cell(i, 15).value) + "," + str(sh.cell(i, 16).value) + "," + str(
                sh.cell(i, 17).value) + "," + str(sh.cell(i, 18).value) + "," + str(sh.cell(i, 20).value).replace("G1","Group/Listed").replace("G2","Group/Listed").replace("G3","Group/Listed").replace("Gri","Griffin")
            key2=str(sh.cell(i, 15).value) + "," + str(sh.cell(i, 16).value) + "," + str(sh.cell(i, 18).value) + "," + str(sh.cell(i, 20).value).replace("G1","Group/Listed").replace("G2","Group/Listed").replace("G3","Group/Listed").replace("Gri","Griffin")

            if key in pawtime:
                val = pawtime[key] / pawCounter[key] if pawCounter[key] > 0 else 0
                a = math.floor(val / 6000)
                val = val % 6000
                b = math.floor(val / 100)
                val = val % 100
                c = math.floor(val)

                sh.cell(i, 22).value = f"{str(a).zfill(2)}.{str(b).zfill(2)}.{str(c).zfill(2)}"
                # partimefull[key] = f"{str(a).zfill(2)}.{str(b).zfill(2)}.{str(c).zfill(2)}"

                if len(str(sh.cell(i, 18).value)) > 0 and not sh.cell(i, 18).value is None and len(key.split(","))==5:
                    # print(sh.cell(i, 35).value,f"{str(a).zfill(2)}.{str(b).zfill(2)}.{str(c).zfill(2)}")
                    # sh.cell(i, 38).value=sh.cell(i, 38).value
                    dist=int(str(sh.cell(i, 18).value).replace("M",""))
                    if str(sh.cell(i, 15).value).replace(" ", "").upper() == "HV":
                        sh.cell(i, 23).value = getTurnValue(hv,  sh.cell(i, 22).value, dist)

                    elif str(sh.cell(i, 16).value).replace(" ", "").upper()=="ST" and len(str(sh.cell(i, 5).value).lower().replace(" ",
                                                                                                              "").replace("none","")) >3 :
                        sh.cell(i, 23).value = getTurnValue(stAwt,  sh.cell(i, 22).value, dist)
                    else:
                        sh.cell(i, 23).value = getTurnValue(st,  sh.cell(i, 22).value, dist)

                pawratings[key] = [f"{str(a).zfill(2)}.{str(b).zfill(2)}.{str(c).zfill(2)}", sh.cell(i, 23).value]
            if key2 in pawtime:
                val = pawtime[key2] / pawCounter[key2] if pawCounter[key2] > 0 else 0
                a = math.floor(val / 6000)
                val = val % 6000
                b = math.floor(val / 100)
                val = val % 100
                c = math.floor(val)

                if len(str(sh.cell(i, 18).value)) > 0 and not sh.cell(i, 18).value is None and len(key.split(",")) == 5:
                    # print(sh.cell(i, 35).value,f"{str(a).zfill(2)}.{str(b).zfill(2)}.{str(c).zfill(2)}")
                    # sh.cell(i, 38).value=sh.cell(i, 38).value
                    dist = int(str(sh.cell(i, 18).value).replace("M", ""))
                    if str(sh.cell(i, 15).value).replace(" ", "").upper() == "HV":
                        sh.cell(i, 23).value = getTurnValue(hv,   f"{str(a).zfill(2)}.{str(b).zfill(2)}.{str(c).zfill(2)}", dist)

                    elif str(sh.cell(i, 16).value).replace(" ", "").upper()=="ST" and len(str(sh.cell(i, 5).value).lower().replace(" ",
                                                                                                              "").replace("none","")) >3 :
                        sh.cell(i, 23).value = getTurnValue(stAwt,   f"{str(a).zfill(2)}.{str(b).zfill(2)}.{str(c).zfill(2)}", dist)
                    else:
                        sh.cell(i, 23).value = getTurnValue(st,   f"{str(a).zfill(2)}.{str(b).zfill(2)}.{str(c).zfill(2)}", dist)

                    pawratings[key2] = [f"{str(a).zfill(2)}.{str(b).zfill(2)}.{str(c).zfill(2)}", val]



        # print( sh.cell(i, 4).value,len(str(sh.cell(i, 4).value).split("/")),not sh.cell(i, 4).value is None and len(str(sh.cell(i, 4).value).split("/"))>2)

        i = 1

    for key in pawtime:
        arr=key.split(",")

        if len(arr) ==5:
            if arr[0].lower()=="hv" and class_Name_Switcher(str(arr[4]).strip()) >=0:

                # print("G",arr[0].lower()+arr[3],classParColumns[arr[0].lower()+arr[3]],str(sp.cell(classParColumns[arr[0].lower()+arr[3]],3).value),"Group"in  str(sp.cell(classParColumns[arr[0].lower()+arr[3]],3).value))
                if "Group"in  str(sp.cell(classParColumns[arr[0].lower()+arr[3]],3).value):
                    column=classParColumns[arr[0].lower()+arr[3]] +  class_Name_Switcher(str(arr[4]).strip())

                else:
                    # print("P", classParColumns[arr[0].lower() + arr[3]] + class_Name_Switcher(str(arr[4]).strip()))

                    column=classParColumns[arr[0].lower() + arr[3]] + class_Name_Switcher(str(arr[4]).strip()) -1

                sp.cell(column, row_Switcher(arr[2])).value = pawratings[key][0]
                sp.cell(column, row_Switcher(arr[2])+1).value = pawratings[key][1]
            elif arr[0].lower() == "st" and class_Name_Switcher(str(arr[4]).strip())>=0 and not "weathe" in arr[1].lower():

                if "Group" in str(sp.cell(classParColumns["sh" + arr[3]], 3).value) :
                    column = classParColumns["sh" + arr[3]] + class_Name_Switcher(str(arr[4]).strip())

                else:
                    column=classParColumns["sh" + arr[3]] + class_Name_Switcher(str(arr[4]).strip()) - 1

                sp.cell(column, row_Switcher(arr[2])).value = pawratings[key][0]
                sp.cell(column, row_Switcher(arr[2])+1).value = pawratings[key][1]
            elif "weather" in arr[1].lower():

                column = classParColumns["awt" + arr[3]] + class_Name_Switcher(str(arr[4]).strip())-1
                sp.cell(column, 4).value = pawratings[key][0]
                # print("awt" + arr[3],column,pawratings[key][0],pawratings[key][1])
                sp.cell(column, 5).value = pawratings[key][1]
        else:
            if arr[0].lower() == "hv" and class_Name_Switcher(str(arr[3]).strip()) >= 0:

                # print(arr[0].lower()+arr[3],classParColumns[arr[0].lower()+arr[3]],str(sp.cell(classParColumns[arr[0].lower()+arr[3]],3).value),"Group"in  str(sp.cell(classParColumns[arr[0].lower()+arr[3]],3).value))
                if "Group" in str(sp.cell(classParColumns[arr[0].lower() + arr[2]], 3).value):
                    column = classParColumns[arr[0].lower() + arr[2]] + class_Name_Switcher(str(arr[3]).strip())

                else:
                    column=classParColumns[arr[0].lower() + arr[2]] + class_Name_Switcher(str(arr[3]).strip()) - 1
                # print(column, class_Name_Switcher(arr[3]), arr[3])
                sp.cell(column, 4).value = pawratings[key][0]
                sp.cell(column, 5).value = pawratings[key][1]
            elif arr[0].lower() == "st" and class_Name_Switcher(str(arr[3]).strip()) >= 0 and not "weathe" in arr[
                1].lower():

                if "Group" in str(sp.cell(classParColumns["sh" + arr[2]], 3).value):
                    column = classParColumns["sh" + arr[2]] + class_Name_Switcher(str(arr[3]).strip())

                else:
                    column=classParColumns["sh" + arr[2]] + class_Name_Switcher(str(arr[3]).strip()) - 1

                sp.cell(column, 4).value = pawratings[key][0]
                sp.cell(column, 5).value = pawratings[key][1]

            # sp.cell(column, 4).value = key
        # print(column)
    saveme(wk, Address)
