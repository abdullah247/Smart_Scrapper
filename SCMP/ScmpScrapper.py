from selenium import webdriver
import openpyxl
from selenium.webdriver.common.by import By
import  datetime as dt
from SCMP.ScmpFunctions import excel_date, saveme
from SCMP.ScmpProcess import getDates






class detailsData:
    def __init__(self):
        mclass=""
        going =""
        course=""
        dist=""
        track=""
        rc=""

def getMData(driver):
    try:
        mdata = detailsData()
        details = str(driver.find_element(By.CSS_SELECTOR,".details>p").text).split("\n")
        for detail in details:
            # print(detail)
            if detail.startswith("Class"):
                mdata.mclass = detail.split("-")[0].replace("Class", "").strip()
                mdata.dist = detail.split("-")[1].strip()
                if len(detail.split()) > 2:
                    mdata.course = detail.split("-")[2].strip()
            elif detail.startswith("Course"):
                mdata.rc = detail.replace("Course:", "")
            elif detail.startswith("Going:"):
                mdata.going = detail.replace("Going:", "")
    except Exception as ex:
        print("get M Data ", str(ex))
    return mdata

    pass


def getMainData(sh,driver,mdate):
    global rowCounter
    try:
        breakCounter = 0
        trs = driver.find_elements(By.CSS_SELECTOR,".race-table tr")
        while len(trs) < 2:
            trs = driver.find_elements(By.CSS_SELECTOR,".race-table tr")
            # time.sleep(1)
            breakCounter += 1
            if breakCounter > 5000:
                break
        trs = driver.find_elements(By.CSS_SELECTOR,".race-table tr")

        mdata = getMData(driver)

        for tr in trs:

            if not str(tr.find_element(By.CSS_SELECTOR,"td").text).strip() == "Place":
                rowCounter += 1
                columnCounter = 0
                dat=mdate.split(",")
                sh.cell(rowCounter, 21).value = excel_date(dt.datetime(int(dat[0]),int(dat[1]),int(dat[2])))
                sh.cell(rowCounter, 20).value = mdata.mclass
                sh.cell(rowCounter, 19).value = mdata.going
                sh.cell(rowCounter, 18).value = mdata.dist
                sh.cell(rowCounter, 15).value = str(mdata.rc).strip().split(" ")[0]

                if len(mdata.course) < 1:
                    sh.cell(rowCounter, 16).value = "turf"
                elif len(mdata.course) <= 4 or "+" in mdata.course:
                    sh.cell(rowCounter, 17).value = mdata.course
                else:
                    sh.cell(rowCounter, 16).value = mdata.course

                for td in tr.find_elements(By.CSS_SELECTOR,"td"):
                    columnCounter += 1
                    sh.cell(rowCounter, columnCounter).value = str(td.text)
                    if len(td.find_elements(By.CSS_SELECTOR,"a"))>0:
                        sh.cell(rowCounter, columnCounter).hyperlink=td.find_elements(By.CSS_SELECTOR,"a")[0].get_attribute("href")
    except Exception as ex:
        print("get main Data " , str(ex))






def getAlllreadyThere(sh,alreadyThere):
    for i in range(2,sh.max_row+1):
        if not sh.cell(i,3).value is None and len(str(sh.cell(i,3).value).strip())>0:
            alreadyThere.append(str(sh.cell(i,3).value)+str(sh.cell(i,21).value))
    pass


def SCARP_SCMP(Address):
    GAMEDATES = []
    alreadyThere = []
    rowCounter = 1
    savecounter=0
    base="https://www.scmp.com/sport/racing/race-result/"
    url = "https://www.scmp.com/sport/racing/race-result"
    print(Address)
    wk = openpyxl.load_workbook(Address)
    sh = wk["Database"]
    options = webdriver.ChromeOptions();
    # options.add_argument('headless');
    options.add_argument('--start-maximized');
    dbcount = 2
    # getAlllreadyThere(sh)
    driver = webdriver.Chrome( options=options)
    driver.get(url)
    driver.refresh()
    GAMEDATES,rowCount=getDates(driver,str(wk["StartDate"].cell(1,1).value))
    rowCounter = int(rowCount) if rowCount=="1" else sh.max_row

    for dat in GAMEDATES:
        # print("working")
        driver.get(base + str(dat).replace(",","") +"/1")
        getMainData(sh,driver,dat)
        races = len(driver.find_elements(By.CSS_SELECTOR,".lists>li"))
        for i in range(1,races):
            driver.get(base + str(dat).replace(",", "") + f"/{i+1}")
            getMainData(sh, driver,dat)

    saveme(wk,Address)
    savecounter=0